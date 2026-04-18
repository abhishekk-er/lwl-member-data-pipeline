"""
Zoho CRM Lead Data Pipeline
============================
Full pipeline: raw CRM export → cleaned Zoho-ready import file

Usage:
    python pipeline.py --raw raw_data.csv --cleaned cleaned_data.csv --output output.xlsx

NOTE: Run on your own data files. Sample data in /sample_data/ uses synthetic records.
"""

import pandas as pd
import numpy as np
import re
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ── Target Zoho CRM columns ───────────────────────────────────────────────────
TARGET_COLS = [
    'First Name', 'Last Name', 'Full Name', 'Email', 'Mobile',
    'CompanyName', 'Industry', 'Designation', 'Website',
    'Years of Experience', 'Tier', 'Lead Type', 'City',
    'Lead Owner', 'Lead Source', 'Lead Status', 'Created By',
    'Modified By', 'Lead Quality', 'Remarks',
    'LinkedIn Profile', 'Instagram Profile'
]


# ═════════════════════════════════════════════════════════════════════════════
# STEP 1 — COLUMN MAPPING
# ═════════════════════════════════════════════════════════════════════════════

def map_columns(df):
    """Map raw 59-column CRM export to 22 Zoho Lead fields."""

    out = pd.DataFrame(index=df.index, columns=TARGET_COLS)

    out['First Name']          = df.get('First Name')
    out['Last Name']           = df.get('Last Name')
    out['Full Name']           = df.get('Contact Name')
    out['Email']               = df.get('Email')
    out['CompanyName']         = df.get('CompanyName')
    out['Industry']            = df.get('Industry')
    out['Designation']         = df.get('Designation')
    out['Website']             = np.nan
    out['Tier']                = df.get('Tier')
    out['Lead Type']           = df.get('Tag')           # renamed
    out['City']                = df.get('City')
    out['Lead Owner']          = df.get('Contact Owner') # renamed
    out['Lead Source']         = df.get('Lead Source')
    out['Lead Status']         = df.get('Membership Status')  # renamed
    out['Created By']          = df.get('Created By')
    out['Modified By']         = df.get('Modified By')
    out['Lead Quality']        = np.nan
    out['Remarks']             = df.get('Description')   # renamed
    out['LinkedIn Profile']    = df.get('LinkedIn Profile')
    out['Instagram Profile']   = df.get('Instagram Profile')

    # Mobile and YoE handled in dedicated steps below
    return out


# ═════════════════════════════════════════════════════════════════════════════
# STEP 2 — MOBILE NUMBER NORMALIZATION
# ═════════════════════════════════════════════════════════════════════════════

def clean_mobile(val):
    """
    Normalize mobile numbers to clean 10-digit Indian format.
    Handles: +91 prefix, scientific notation, dashes, spaces, commas.
    """
    if pd.isna(val) or str(val).strip() in ('', 'nan'):
        return np.nan
    s = str(val).strip()
    digits = re.sub(r'\D', '', s)
    if len(digits) > 10 and digits.startswith('91'):
        digits = digits[2:]
    if len(digits) > 10:
        digits = digits[-10:]
    return digits if len(digits) >= 8 else s  # keep original if suspicious


def apply_mobile_cleaning(df, raw_df):
    """Use Mobile final if available, fall back to Mobile. Then clean."""
    mob_final = raw_df.get('Mobile final', pd.Series(dtype=str)).apply(clean_mobile)
    mob_orig  = raw_df.get('Mobile', pd.Series(dtype=str)).apply(clean_mobile)
    df['Mobile'] = mob_final.where(mob_final.notna(), mob_orig)
    return df


# ═════════════════════════════════════════════════════════════════════════════
# STEP 3 — YEARS OF EXPERIENCE MERGE
# ═════════════════════════════════════════════════════════════════════════════

INVALID_YOE = {'select', 'event guest', 'nan', '', 'none'}
INCOME_KEYWORDS = ['lakh', 'lacs', 'lac', 'crore', 'aed', 'inr', '₹']

def parse_yoe_text(val):
    """Extract numeric YoE from text like '5 to 10 years', '15+', '1 - 5 years'."""
    if pd.isna(val): return np.nan
    s = str(val).strip()
    if s.lower() in INVALID_YOE: return np.nan
    if any(k in s.lower() for k in INCOME_KEYWORDS): return np.nan
    nums = re.findall(r'\d+', s)
    return int(nums[0]) if nums else np.nan


def merge_yoe_columns(df, raw_df):
    """Merge numeric and text YoE columns into one."""
    yoe_num  = pd.to_numeric(raw_df.get('Years of Experience', pd.Series()), errors='coerce')
    yoe_text = raw_df.get('Years of Experience -1', pd.Series(dtype=str)).apply(parse_yoe_text)
    merged   = yoe_num.where(yoe_num.notna(), yoe_text)
    df['Years of Experience'] = merged.apply(
        lambda v: int(v) if pd.notna(v) and float(v) == int(float(v)) else (v if pd.notna(v) else np.nan)
    )
    return df


# ═════════════════════════════════════════════════════════════════════════════
# STEP 4 — FIRST / LAST NAME EXTRACTION
# ═════════════════════════════════════════════════════════════════════════════

def is_blank(val):
    return pd.isna(val) or str(val).strip() in ('', 'nan', 'NaN', '.', 'NaT', 'None')


def fix_names(df):
    """
    For rows missing First Name, extract from Full Name.
    Zoho rule: single-word names go in Last Name only (Last Name is mandatory).
    """
    for idx, row in df.iterrows():
        if not is_blank(row['First Name']):
            continue

        source = None
        if not is_blank(row['Full Name']):
            source = str(row['Full Name']).strip().rstrip('.').strip()
        elif not is_blank(row['Last Name']) and str(row['Last Name']).strip() not in ('.', ''):
            source = str(row['Last Name']).strip().rstrip('.').strip()

        if not source:
            continue

        parts = [p for p in source.split() if p]
        if len(parts) == 1:
            df.at[idx, 'First Name'] = np.nan      # blank — Zoho only needs Last Name
            df.at[idx, 'Last Name']  = parts[0]
        elif len(parts) >= 2:
            df.at[idx, 'First Name'] = parts[0]
            df.at[idx, 'Last Name']  = ' '.join(parts[1:])

    # Clean stray dots from Last Name
    df['Last Name'] = df['Last Name'].apply(
        lambda v: np.nan if str(v).strip() in ('.', 'nan', 'NaN', '', 'None') else v
    )
    return df


# ═════════════════════════════════════════════════════════════════════════════
# STEP 5 — YOE ENRICHMENT (Cleaned ← Raw, email join)
# ═════════════════════════════════════════════════════════════════════════════

def is_richer_yoe(val):
    """True if value is a descriptive experience range (not a plain number or income value)."""
    if pd.isna(val) or str(val).strip().lower() in INVALID_YOE:
        return False
    s = str(val).strip()
    if any(k in s.lower() for k in INCOME_KEYWORDS):
        return False
    return bool(re.search(r'(year|yr|above|to|\+|-)', s, re.IGNORECASE))


def enrich_yoe(cleaned_df, raw_df):
    """
    Replace plain numeric YoE in cleaned data with richer text from raw,
    matched by email address.
    """
    cleaned_df['Years of Experience'] = cleaned_df['Years of Experience'].astype(object)

    # Build email → richer YoE lookup from raw
    raw_df['_key'] = raw_df['Email'].astype(str).str.strip().str.lower()
    yoe_map = {}
    for _, row in raw_df.iterrows():
        key = row['_key']
        if key in ('nan', ''): continue
        y2 = row.get('Years of Experience -1')
        y1 = row.get('Years of Experience')
        if is_richer_yoe(y2):
            yoe_map[key] = str(y2).strip()
        elif is_richer_yoe(y1):
            yoe_map[key] = str(y1).strip()

    # Apply
    cleaned_df['_key'] = cleaned_df['Email'].astype(str).str.strip().str.lower()
    replaced = 0
    for idx, row in cleaned_df.iterrows():
        richer = yoe_map.get(row['_key'])
        if richer:
            cleaned_df.at[idx, 'Years of Experience'] = richer
            replaced += 1

    cleaned_df.drop(columns=['_key'], inplace=True)
    print(f"  YoE enriched: {replaced} rows updated")
    return cleaned_df


# ═════════════════════════════════════════════════════════════════════════════
# OUTPUT — Plain white Excel (Google Sheets compatible)
# ═════════════════════════════════════════════════════════════════════════════

def save_to_excel(df, output_path):
    cols = list(df.columns)
    wb = Workbook()
    ws = wb.active
    ws.title = "Zoho Leads"

    hdr_font = Font(name='Arial', bold=True, size=10)
    dat_font = Font(name='Arial', size=10)
    align    = Alignment(vertical='center')

    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = hdr_font; c.alignment = align

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            v = None if (isinstance(val, float) and np.isnan(val)) else val
            v = None if str(v).strip() in ('nan', 'NaN', 'NaT', 'None', '') else v
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = dat_font; c.alignment = align

    col_widths = {
        'First Name': 14, 'Last Name': 14, 'Full Name': 22, 'Email': 32,
        'Mobile': 14, 'CompanyName': 28, 'Industry': 20, 'Designation': 24,
        'Website': 22, 'Years of Experience': 22, 'Tier': 8, 'Lead Type': 14,
        'City': 14, 'Lead Owner': 20, 'Lead Source': 14, 'Lead Status': 14,
        'Created By': 18, 'Modified By': 18, 'Lead Quality': 12, 'Remarks': 28,
        'LinkedIn Profile': 30, 'Instagram Profile': 28
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 16)

    ws.freeze_panes = 'A2'
    wb.save(output_path)
    print(f"  Saved → {output_path}  ({len(df):,} rows | {len(cols)} cols)")


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════

def run_pipeline(raw_path, output_path):
    print("=" * 55)
    print("  Zoho CRM Lead Data Pipeline")
    print("=" * 55)

    print("\n[1/5] Loading raw data...")
    raw_df = pd.read_csv(raw_path, low_memory=False) if raw_path.endswith('.csv') \
             else pd.read_excel(raw_path)
    print(f"  Raw data: {raw_df.shape[0]:,} rows × {raw_df.shape[1]} cols")

    print("\n[2/5] Mapping columns...")
    cleaned = map_columns(raw_df)
    cleaned = apply_mobile_cleaning(cleaned, raw_df)
    cleaned = merge_yoe_columns(cleaned, raw_df)
    print(f"  Mapped to {len(TARGET_COLS)} Zoho fields")

    print("\n[3/5] Cleaning string values...")
    for col in cleaned.select_dtypes(include='object').columns:
        cleaned[col] = cleaned[col].astype(str).str.strip()
        cleaned[col] = cleaned[col].replace({'nan': np.nan, 'NaT': np.nan, 'None': np.nan, '.': np.nan})

    print("\n[4/5] Fixing names (First/Last from Full Name)...")
    cleaned = fix_names(cleaned)
    missing_last = cleaned['Last Name'].isna().sum()
    print(f"  Still missing Last Name (no source): {missing_last}")

    print("\n[5/5] Saving output...")
    save_to_excel(cleaned, output_path)

    print("\n✅ Pipeline complete!")
    print("=" * 55)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Zoho CRM Lead Data Pipeline')
    parser.add_argument('--raw',     required=True,  help='Path to raw data file (.csv or .xlsx)')
    parser.add_argument('--output',  required=True,  help='Output file path (.xlsx)')
    args = parser.parse_args()
    run_pipeline(args.raw, args.output)
